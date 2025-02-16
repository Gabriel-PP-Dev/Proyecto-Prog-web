from django.shortcuts import render
from rest_framework import viewsets
from rest_framework.response import Response
from rest_framework.decorators import action
from .models import Usuario, Tienda, Producto, TiendaProductoPrecio, Producto_Precio, Venta
from .serializers import UsuarioSerializer, TiendaSerializer, ProductoSerializer, TiendaProductoPrecioSerializer, Producto_PrecioSerializer, VentaSerializer
from datetime import datetime
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from io import BytesIO
from django.db.models import Sum
from drf_yasg.utils import swagger_auto_schema
from drf_yasg import openapi
from rest_framework.permissions import IsAuthenticated
from rest_framework.permissions import AllowAny
from rest_framework import status
from openpyxl import Workbook
from openpyxl.styles import Font

# Create your views here.

class UsuarioViewSet(viewsets.ModelViewSet):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioSerializer
    #permission_classes = [IsAuthenticated]

    @action(methods=['post'], detail=False, permission_classes=[AllowAny])
    @swagger_auto_schema(
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=['nombre_usuario', 'contrasenna'],
            properties={
                'nombre_usuario': openapi.Schema(type=openapi.TYPE_STRING, description='Nombre de usuario'),
                'contrasenna': openapi.Schema(type=openapi.TYPE_STRING, description='Contraseña')
            }
        ),
        responses={
            200: openapi.Response('Usuario autenticado correctamente', UsuarioSerializer),
            401: openapi.Response('Credenciales incorrectas'),
            404: openapi.Response('Usuario no encontrado')
        }
    )
    def login(self, request):
        data = request.data
        nombre_usuario = data.get('nombre_usuario')
        contrasenna = data.get('contrasenna')

        if not nombre_usuario or not contrasenna:
            return Response({'error': 'Faltan credenciales'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            usuario = Usuario.objects.get(nombre_usuario=nombre_usuario)
        except Usuario.DoesNotExist:
            return Response({'error': 'Usuario no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        if not usuario.check_password(contrasenna):
            return Response({'error': 'Credenciales incorrectas'}, status=status.HTTP_401_UNAUTHORIZED)

        serializer = UsuarioSerializer(usuario)
        return Response(serializer.data)
    
    @action(methods=['get'], detail=False)
    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                'nombre',
                openapi.IN_PATH,
                description='Nombre del usuario',
                type=openapi.TYPE_STRING,
                required=True
            )
        ],
        responses={
            200: openapi.Response('Usuarios encontrados', UsuarioSerializer(many=True))
        }
    )
    def buscar_por_nombre(self, request, *args, **kwargs):
        nombre = kwargs['nombre']
        usuarios = Usuario.objects.filter(nombre__icontains=nombre)
        serializer = UsuarioSerializer(usuarios, many=True)
        return Response(serializer.data)

class TiendaViewSet(viewsets.ModelViewSet):
    queryset = Tienda.objects.all()
    serializer_class = TiendaSerializer
    #permission_classes = [IsAuthenticated]

    @action(methods=['get'], detail=False)
    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                'nombre',
                openapi.IN_PATH,
                description='Nombre de la tienda',
                type=openapi.TYPE_STRING,
                required=True
            )
        ],
        responses={
            200: openapi.Response('Tiendas encontradas', TiendaSerializer(many=True))
        }
    )
    def buscar_por_nombre(self, request, *args, **kwargs):
        nombre = kwargs['nombre']
        tiendas = Tienda.objects.filter(nombre__icontains=nombre)
        serializer = TiendaSerializer(tiendas, many=True)
        return Response(serializer.data)

class ProductoViewSet(viewsets.ModelViewSet):
    queryset = Producto.objects.all()
    serializer_class = ProductoSerializer
    #permission_classes = [IsAuthenticated]
    

class TiendaProductoPrecioViewSet(viewsets.ModelViewSet):
    queryset = TiendaProductoPrecio.objects.all()
    serializer_class = TiendaProductoPrecioSerializer
    #permission_classes = [IsAuthenticated]

    @action(methods=['get'], detail=False)
    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                'producto_id',
                openapi.IN_PATH,
                description='ID del producto',
                type=openapi.TYPE_STRING,
                required=True
            )
        ],
        responses={
            200: openapi.Response('Lista de TiendaProductoPrecio relacionados con el producto', TiendaProductoPrecioSerializer(many=True)),
            404: openapi.Response('Producto no encontrado')
        }
    )
    def producto_por_id(self, request, producto_id):
        try:
            producto = Producto.objects.get(id=producto_id)
        except Producto.DoesNotExist:
            return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        tienda_producto_precio = TiendaProductoPrecio.objects.filter(producto_precio__producto=producto)
        serializer = TiendaProductoPrecioSerializer(tienda_producto_precio, many=True)
        return Response(serializer.data)
    

    @action(methods=['get'], detail=False)
    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                'tienda_id',
                openapi.IN_PATH,
                description='ID de la tienda',
                type=openapi.TYPE_STRING,
                required=True
            )
        ],
        responses={
            200: openapi.Response('Lista de TiendaProductoPrecio relacionados con la tienda', TiendaProductoPrecioSerializer(many=True)),
            404: openapi.Response('Tienda no encontrada')
        }
    )
    def tienda_por_id(self, request, tienda_id):
        try:
            tienda = Tienda.objects.get(id=tienda_id)
        except Tienda.DoesNotExist:
            return Response({'error': 'Tienda no encontrada'}, status=status.HTTP_404_NOT_FOUND)

        tienda_producto_precio = TiendaProductoPrecio.objects.filter(tienda=tienda)
        serializer = TiendaProductoPrecioSerializer(tienda_producto_precio, many=True)
        return Response(serializer.data)
    

    @action(methods=['get'], detail=False)
    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                'tienda_id',
                openapi.IN_PATH,
                description='ID de la tienda',
                type=openapi.TYPE_STRING,
                required=True
            )
        ],
        responses={
            200: openapi.Response('Lista de TiendaProductoPrecio relacionados con la tienda, ordenados de mayor a menor', TiendaProductoPrecioSerializer(many=True)),
            404: openapi.Response('Tienda no encontrada')
        }
    )
    def tienda_por_id_ordenado(self, request, tienda_id):
        try:
            tienda = Tienda.objects.get(id=tienda_id)
        except Tienda.DoesNotExist:
            return Response({'error': 'Tienda no encontrada'}, status=status.HTTP_404_NOT_FOUND)

        tienda_producto_precio = TiendaProductoPrecio.objects.filter(tienda=tienda).order_by('-cantidad_en_tienda')
        serializer = TiendaProductoPrecioSerializer(tienda_producto_precio, many=True)
        return Response(serializer.data)
    
    @action(methods=['post'], detail=False)
    @swagger_auto_schema(
        request_body=openapi.Schema(
            type=openapi.TYPE_OBJECT,
            required=['producto_precio', 'tienda_origen', 'tienda_destino', 'cantidad'],
            properties={
                'producto_precio': openapi.Schema(type=openapi.TYPE_STRING, description='ID del producto precio'),
                'tienda_origen': openapi.Schema(type=openapi.TYPE_STRING, description='ID de la tienda origen'),
                'tienda_destino': openapi.Schema(type=openapi.TYPE_STRING, description='ID de la tienda destino'),
                'cantidad': openapi.Schema(type=openapi.TYPE_INTEGER, description='Cantidad a mover')
            }
        ),
        responses={
            200: openapi.Response('Movimiento realizado con éxito'),
            400: openapi.Response('Error en la solicitud'),
            404: openapi.Response('Tienda o producto precio no encontrado')
        }
    )
    def mover_producto(self, request):
        data = request.data
        producto_precio_id = data.get('producto_precio')
        tienda_origen_id = data.get('tienda_origen')
        tienda_destino_id = data.get('tienda_destino')
        cantidad = data.get('cantidad')

        if not all([producto_precio_id, tienda_origen_id, tienda_destino_id, cantidad]):
            return Response({'error': 'Faltan datos en la solicitud'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            producto_precio = Producto_Precio.objects.get(id=producto_precio_id)
            tienda_origen = Tienda.objects.get(id=tienda_origen_id)
            tienda_destino = Tienda.objects.get(id=tienda_destino_id)
        except (Producto_Precio.DoesNotExist, Tienda.DoesNotExist):
            return Response({'error': 'Tienda o producto precio no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        tienda_producto_precio_origen = TiendaProductoPrecio.objects.filter(tienda=tienda_origen, producto_precio=producto_precio).first()
        if not tienda_producto_precio_origen:
            return Response({'error': 'El producto precio no se encuentra en la tienda origen'}, status=status.HTTP_400_BAD_REQUEST)

        if tienda_producto_precio_origen.cantidad_en_tienda < cantidad:
            return Response({'error': 'No hay cantidad suficiente para realizar el movimiento'}, status=status.HTTP_400_BAD_REQUEST)

        tienda_producto_precio_destino = TiendaProductoPrecio.objects.filter(tienda=tienda_destino, producto_precio=producto_precio).first()
        if tienda_producto_precio_destino:
            tienda_producto_precio_destino.cantidad_en_tienda += cantidad
            tienda_producto_precio_destino.save()
        else:
            TiendaProductoPrecio.objects.create(tienda=tienda_destino, producto_precio=producto_precio, cantidad_en_tienda=cantidad)

        tienda_producto_precio_origen.cantidad_en_tienda -= cantidad
        tienda_producto_precio_origen.save()

        return Response({'mensaje': 'Movimiento realizado con éxito'})
class Producto_PrecioViewSet(viewsets.ModelViewSet):
    queryset = Producto_Precio.objects.all()
    serializer_class = Producto_PrecioSerializer
    #permission_classes = [IsAuthenticated]

    @action(methods=['get'], detail=False, permission_classes=[AllowAny])
    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                'producto_id',
                openapi.IN_PATH,
                description='ID del producto',
                type=openapi.TYPE_STRING,
                required=True
            )
        ],
        responses={
            200: openapi.Response('Lista de Producto_Precio relacionados con el producto', Producto_PrecioSerializer(many=True)),
            404: openapi.Response('Producto no encontrado')
        }
    )
    def producto_precio_por_producto(self, request, producto_id):
        try:
            producto = Producto.objects.get(id=producto_id)
        except Producto.DoesNotExist:
            return Response({'error': 'Producto no encontrado'}, status=status.HTTP_404_NOT_FOUND)

        producto_precio = Producto_Precio.objects.filter(producto=producto)
        serializer = Producto_PrecioSerializer(producto_precio, many=True)
        return Response(serializer.data)

class VentaViewSet(viewsets.ModelViewSet):
    queryset = Venta.objects.all()
    serializer_class = VentaSerializer
    #permission_classes = [IsAuthenticated]

    @action(methods=['get'], detail=False)
    @swagger_auto_schema(
        manual_parameters=[
            openapi.Parameter(
                'tienda_id',
                openapi.IN_PATH,
                description='ID de la tienda',
                type=openapi.TYPE_STRING,
                required=True
            )
        ],
        responses={
            200: openapi.Response('Lista de ventas relacionadas con la tienda', VentaSerializer(many=True)),
            404: openapi.Response('Tienda no encontrada')
        }
    )
    def ventas_por_tienda(self, request, tienda_id):
        try:
            tienda = Tienda.objects.get(id=tienda_id)
        except Tienda.DoesNotExist:
            return Response({'error': 'Tienda no encontrada'}, status=status.HTTP_404_NOT_FOUND)

        ventas = Venta.objects.filter(tienda_producto_precio__tienda=tienda)
        serializer = VentaSerializer(ventas, many=True)
        return Response(serializer.data)
    
    @action(methods=['get'], detail=False)
    def exportar_ventas(self, request):
        ventas = Venta.objects.all()
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas"

        # Establecer encabezados
        ws['A1'] = "Cantidad"
        ws['B1'] = "Precio"
        ws['C1'] = "ID TiendaProductoPrecio"
        ws['D1'] = "ID Venta"

        # Establecer estilo para encabezados
        for cell in ws["1:1"][0]:
            cell.font = Font(bold=True)

        # Agregar datos
        for i, venta in enumerate(ventas, start=2):
            ws[f'A{i}'] = venta.cantidad
            ws[f'B{i}'] = venta.precio
            ws[f'C{i}'] = venta.tienda_producto_precio.id
            ws[f'D{i}'] = venta.id

        # Establecer nombre del archivo
        filename = "ventas.xlsx"

        # Establecer respuesta HTTP
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        # Guardar archivo en respuesta HTTP
        wb.save(response)

        return response
